import os
import numpy as np
import random
# import cPickle as pickle
import _pickle as pickle
from scipy.io import loadmat

import openpyxl

np.random.seed(0)
random.seed(0)

def make_dir(path):
    if os.path.exists(path):
        pass
    else:
        os.mkdir(path)

def generate_avtech_attr_value(origin_attr_gt, index):
    return 1 if any(origin_attr_gt[idx] == 1 for idx in index) else 0

def generate_data_description(save_dir):
    """
    create a dataset description file, which consists of images, labels
    """
    dataset = dict()
    dataset['description'] = 'peta'
    dataset['root'] = '../Dataset/PETA/images/'
    dataset['image'] = []
    dataset['att'] = []
    dataset['att_name'] = []
    dataset['selected_attribute'] = [2, 3, 4, 10, 16, 17, 30, 35, 36, 37, 38, 43, 44, 45, 46, 47, 48, 49, 54, 55, 56, 80, 87, 105, 106, 107, 108]
    # dataset['selected_attribute'] = [93, 106, 107, 108]

    dataset['test_root'] = '../Dataset/rap2/RAP_dataset/'
    dataset['test_image'] = []

    # load PETA.MAT
    data = loadmat('../Dataset/PETA/PETA.mat')
    # data = loadmat(open('./Dataset/PETA/PETA.mat', 'r'))  #   Modified
    for idx_i in range(105):
        dataset['att_name'].append(data['peta'][0][0][1][idx_i,0][0])
    dataset['att_name'].extend(['adult', 'upperBodyShort', 'lowerBodyShort', 'lowerBodyLong'])
    corr_attr = [
        [0, 1],
        [26, 97],
        [25, 27, 90],
        [12, 31, 84, 92, 102]
    ]

    for idx_i in range(19000):
        dataset['image'].append('%05d.png'%(idx_i+1))
        dataset['att'].append(data['peta'][0][0][0][idx_i, 4:].tolist())
        dataset['att'][idx_i].extend(np.zeros(4).tolist())

        for idx_j in range(2):
            dataset['att'][idx_i][105 + idx_j] = generate_avtech_attr_value(dataset['att'][idx_i], corr_attr[idx_j])

        if dataset['att'][idx_i][25] == 1:
            dataset['att'][idx_i][107] = 1
        elif dataset['att'][idx_i][27] == 1 or dataset['att'][idx_i][90] == 1:
            dataset['att'][idx_i][107] = 1
        elif dataset['att'][idx_i][12] == 1 or dataset['att'][idx_i][31] == 1 or dataset['att'][idx_i][84] == 1\
             or dataset['att'][idx_i][92] == 1 or dataset['att'][idx_i][102] == 1:
            dataset['att'][idx_i][108] = 1
    
    data = loadmat(open('../Dataset/rap2/RAP_annotation/RAP_annotation.mat', 'rb'))
    for idx in range(84928):
        dataset['test_image'].append(data['RAP_annotation'][0][0][0][idx][0][0])

    ''' Create the excel to record the ground truth '''
    # wb = openpyxl.load_workbook('record_attr.xlsx')
    # s1 = wb['Total_Attr']
    # s2 = wb['UpperBody_Attr']
    # s3 = wb['LowerBody_Attr']

    # ub_attr = [106, 97, 26, 93, 32, 11, 100, 103]
    # lb_attr = [107, 108, 25, 27, 90, 12, 31, 84, 92, 102]

    # for x in range(19001):
    #     for y in range(110):
    #         if x == 0 and y == 0:
    #             s1.cell(x + 1, y + 1).value = 'images'
    #         elif x == 0:
    #             s1.cell(x + 1, y + 1).value = dataset['att_name'][y - 1]
    #         elif y == 0:
    #             s1.cell(x + 1, y + 1).value = dataset['image'][x - 1]
    #         else:
    #             s1.cell(x + 1, y + 1).value = dataset['att'][x - 1][y - 1]
    #     for y in range(len(ub_attr) + 1):
    #         if x == 0 and y == 0:
    #             s2.cell(x + 1, y + 1).value = 'images'
    #         elif x == 0:
    #             s2.cell(x + 1, y + 1).value = dataset['att_name'][ub_attr[y - 1]]
    #         elif y == 0:
    #             s2.cell(x + 1, y + 1).value = dataset['image'][x - 1]
    #         else:
    #             s2.cell(x + 1, y + 1). value = dataset['att'][x - 1][ub_attr[y - 1]]
    #     for y in range(len(lb_attr) + 1):
    #         if x == 0 and y == 0:
    #             s3.cell(x + 1, y + 1).value = 'images'
    #         elif x == 0:
    #             s3.cell(x + 1, y + 1).value = dataset['att_name'][lb_attr[y - 1]]
    #         elif y == 0:
    #             s3.cell(x + 1, y + 1).value = dataset['image'][x - 1]
    #         else:
    #             s3.cell(x + 1, y + 1). value = dataset['att'][x - 1][lb_attr[y - 1]]

    # wb.save('record_attr.xlsx')

    with open(os.path.join(save_dir, 'peta_dataset.pkl'), 'wb+') as f:
        pickle.dump(dataset, f)

def create_trainvaltest_split(traintest_split_file):
    """
    create a dataset split file, which consists of index of the train/val/test splits
    """
    partition = dict()
    partition['trainval'] = []
    partition['train'] = []
    partition['val'] = []
    partition['test'] = []
    partition['weight_trainval'] = []
    partition['weight_train'] = []
    # load PETA.MAT
    data = loadmat(open('../Dataset/PETA/PETA.mat', 'rb'))
    for idx in range(5):
        train = (data['peta'][0][0][3][idx][0][0][0][0][:,0]-1).tolist()
        val = (data['peta'][0][0][3][idx][0][0][0][1][:,0]-1).tolist()
        test = (data['peta'][0][0][3][idx][0][0][0][2][:,0]-1).tolist()
        trainval = train + val
        partition['train'].append(train)
        partition['val'].append(val)
        partition['trainval'].append(trainval)
        partition['test'].append(test)
        
        # weight
        weight_trainval = np.mean(data['peta'][0][0][0][trainval, 4:].astype('float32')==1, axis=0).tolist()
        weight_train = np.mean(data['peta'][0][0][0][train, 4:].astype('float32')==1, axis=0).tolist()

        ''' Calculate the mean value of the new attr '''
        corr_attr = [
        [0, 1],
        [26, 97],
        [25, 27, 90],
        [12, 31, 84, 92, 102]
        ]
        corr_weight_trainval = [[weight_trainval[i] for i in indices] for indices in corr_attr]
        corr_weight_train = [[weight_trainval[i] for i in indices] for indices in corr_attr]
        newAttr_weight_trainval = [np.mean(idx) for idx in corr_weight_trainval]
        newAttr_weight_train = [np.mean(idx) for idx in corr_weight_train]
        weight_trainval.extend(newAttr_weight_trainval)
        weight_train.extend(newAttr_weight_train)

        partition['weight_trainval'].append(weight_trainval)
        partition['weight_train'].append(weight_train)

        partition['weight_trainval'].append(weight_trainval)

    with open(traintest_split_file, 'wb+') as f:
        pickle.dump(partition, f)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="peta dataset")
    parser.add_argument(
        '--save_dir',
        type=str,
        default='../Dataset/PETA/')
    parser.add_argument(
        '--traintest_split_file',
        type=str,
        default="../Dataset/PETA/peta_partition.pkl")
    args = parser.parse_args()
    save_dir = args.save_dir
    traintest_split_file = args.traintest_split_file

    generate_data_description(save_dir)
    create_trainvaltest_split(traintest_split_file)
