import os
import numpy as np
import random
import _pickle as pickle
from scipy.io import loadmat

import openpyxl

np.random.seed(0)
random.seed(0)

def make_dir(path):
    if os.path.exists(path):
        pass
    else:
        os.mkdir(path)

def generate_data_description(save_dir):
    """
    create a dataset description file, which consists of images, labels
    """
    select_att = [0, 1, 4, 5, 16, 31, 32, 34, 35, 36, 38, 39, 53, 54, 56, 57, 58, 60, 61, 88, 89, 152, 153, 154, 155, 156, 157, 158]

    dataset = dict()
    dataset['description'] = 'rap2'
    dataset['root'] = '../Dataset/rap2/RAP_dataset/'
    dataset['image'] = []
    dataset['att'] = []
    dataset['att_name'] = []

    # load RAP_annotation.mat
    data = loadmat(open('../Dataset/rap2/RAP_annotation/RAP_annotation.mat', 'rb'))
    with open('../Dataset/rap2/rap2_extra.pkl', 'rb') as f:
        extra_data = pickle.load(f)

    dataset['selected_attribute'] = select_att
    print(dataset['selected_attribute'])
    
    for idx in range(152):
        dataset['att_name'].append(data['RAP_annotation'][0][0][2][idx][0][0])
    avtech_attr = ['male', 'adult', 'glass', 'upperBodyShort', 'upperBodyLong', 'lowerBodyShort', 'lowerBodyLong']
    dataset['att_name'].extend(avtech_attr)

    ub_short = [23, 24, 29]
    ub_long = [21, 22, 25, 26, 27, 28]
    lb_short = [46, 48]
    lb_long = [45, 49, 50, 51, 52]
    count = 0
    
    for idx in range(84928):
        dataset['image'].append(data['RAP_annotation'][0][0][0][idx][0][0])
        dataset['att'].append(data['RAP_annotation'][0][0][1][idx, :].tolist())
        dataset['att'][idx].extend(np.zeros(len(avtech_attr)).tolist())

        if dataset['att'][idx][0] == 0:
            dataset['att'][idx][152] = 1

        if (dataset['att'][idx][2] == 1) or (dataset['att'][idx][3] == 1):
            dataset['att'][idx][153] = 1

        if (dataset['att'][idx][17] == 1) or (dataset['att'][idx][18] == 1):
            dataset['att'][idx][154] = 1

        if any(dataset['att'][idx][i] == 1 for i in ub_long):
            dataset['att'][idx][156] = 1
        elif extra_data[idx][1] == 1:
            dataset['att'][idx][156] = 1
        elif extra_data[idx][0] == 1:
            dataset['att'][idx][155] = 1

        if any(dataset['att'][idx][i] == 1 for i in lb_long):
            dataset['att'][idx][158] = 1
        elif any(dataset['att'][idx][i] == 1 for i in lb_short) and all(dataset['att'][idx][j] != 1 for j in lb_long):
            dataset['att'][idx][157] = 1

    with open(os.path.join(save_dir, 'rap2_dataset.pkl'), 'wb+') as f:
        pickle.dump(dataset, f)

    ''' Create the excel to record the ground truth '''
    # wb = openpyxl.Workbook()
    # wb.create_sheet('Total_Attr')
    # wb.create_sheet('UpperBody_Attr')
    # wb.create_sheet('LowerBody_Attr')
    # wb.save('rapv2.xlsx')

    # wb = openpyxl.load_workbook('rapv2.xlsx')

    # s1 = wb['Total_Attr']
    # s2 = wb['UpperBody_Attr']
    # s3 = wb['LowerBody_Attr']

    # ub_attr = range(21, 31)
    # lb_attr = range(45, 53)

    # for x in range(84929):
    #     for y in range(153):
    #         if x == 0 and y == 0:
    #             s1.cell(x + 1, y + 1).value = 'images'
    #         elif x == 0:
    #             s1.cell(x + 1, y + 1).value = dataset['att_name'][y - 1]
    #         elif y == 0:
    #             s1.cell(x + 1, y + 1).value = dataset['image'][x - 1]
    #         else:
    #             s1.cell(x + 1, y + 1). value = dataset['att'][x - 1][y - 1]
        # for y in range(len(ub_attr) + 1):
        #     if x == 0 and y == 0:
        #         s2.cell(x + 1, y + 1).value = 'images'
        #     elif x == 0:
        #         s2.cell(x + 1, y + 1).value = dataset['att_name'][ub_attr[y - 1]]
        #     elif y == 0:
        #         s2.cell(x + 1, y + 1).value = dataset['image'][x - 1]
        #     else:
        #         s2.cell(x + 1, y + 1). value = dataset['att'][x - 1][ub_attr[y - 1]]
        # for y in range(len(lb_attr) + 1):
        #     if x == 0 and y == 0:
        #         s3.cell(x + 1, y + 1).value = 'images'
        #     elif x == 0:
        #         s3.cell(x + 1, y + 1).value = dataset['att_name'][lb_attr[y - 1]]
        #     elif y == 0:
        #         s3.cell(x + 1, y + 1).value = dataset['image'][x - 1]
        #     else:
        #         s3.cell(x + 1, y + 1). value = dataset['att'][x - 1][lb_attr[y - 1]]

    # wb.save('rapv2.xlsx')

def create_trainvaltest_split(traintest_split_file):
    """
    create a dataset split file, which consists of index of the train/val/test splits
    """
    partition = dict()
    partition['train'] = []
    partition['val'] = []
    partition['trainval'] = []
    partition['test'] = []
    partition['ub_test'] = []
    partition['weight_train'] = []
    partition['weight_trainval'] = []
    # load RAP_annotation.mat
    data = loadmat(open('../Dataset/rap2/RAP_annotation/RAP_annotation.mat', 'rb'))
    for idx in range(5):
        train = (data['RAP_annotation'][0][0][4][0, idx][0][0][0][0,:]-1).tolist()
        val = (data['RAP_annotation'][0][0][4][0, idx][0][0][1][0,:]-1).tolist()
        test = (data['RAP_annotation'][0][0][4][0, idx][0][0][2][0,:]-1).tolist()
        trainval = train + val
        partition['trainval'].append(trainval)
        partition['train'].append(train)
        partition['val'].append(val)
        partition['test'].append(test)

        ub_test = range(84928)
        partition['ub_test'].append(ub_test)

        # weight
        weight_train = np.mean(data['RAP_annotation'][0][0][1][train, :].astype('float32')==1, axis=0).tolist()
        weight_trainval = np.mean(data['RAP_annotation'][0][0][1][trainval, :].astype('float32')==1, axis=0).tolist()

        ''' Calculate the mean value of the new attr '''
        corr_attr = [
        [0],                    # male
        [2, 3],                 # adult
        [17, 18],               # glass
        [24, 29],
        [21, 22, 25, 26, 27, 28],
        [46, 48],
        [45, 49, 50, 51, 52]
        ]
        corr_weight_trainval = [[weight_trainval[i] for i in indices] for indices in corr_attr]
        corr_weight_train = [[weight_trainval[i] for i in indices] for indices in corr_attr]
        newAttr_weight_trainval = [np.mean(idx) for idx in corr_weight_trainval]
        newAttr_weight_train = [np.mean(idx) for idx in corr_weight_train]
        weight_trainval.extend(newAttr_weight_trainval)
        weight_train.extend(newAttr_weight_train)

        partition['weight_train'].append(weight_train)
        partition['weight_trainval'].append(weight_trainval)

    with open(traintest_split_file, 'wb+') as f:
        pickle.dump(partition, f)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="rap2 dataset")
    parser.add_argument(
        '--save_dir',
        type=str,
        default='../Dataset/rap2/')
    parser.add_argument(
        '--traintest_split_file',
        type=str,
        default="../Dataset/rap2/rap2_partition.pkl")
    args = parser.parse_args()
    save_dir = args.save_dir
    traintest_split_file = args.traintest_split_file

    generate_data_description(save_dir)
    create_trainvaltest_split(traintest_split_file)
